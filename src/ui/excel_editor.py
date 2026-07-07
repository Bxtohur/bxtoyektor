"""Editor Excel sederhana di dalam aplikasi (buka & edit lewat controller).

Menampilkan setiap sheet sebagai tabel yang bisa diedit (QTableWidget dalam
QTabWidget), bisa tambah/hapus baris & kolom, lalu menyimpan kembali ke berkas
.xlsx. Nilai disimpan sebagai teks (aman untuk daftar nama/link; tidak mengutak-
atik tipe angka/tanggal secara diam-diam).

Berkas Excel sumber data bisa langsung diedit di sini, lalu daftar di operator
di-refresh setelah disimpan.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QDialog,
    QHBoxLayout,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QVBoxLayout,
)


class ExcelEditorDialog(QDialog):
    """Dialog non-modal untuk mengedit satu berkas .xlsx."""

    tersimpan = Signal(str)  # path yang berhasil disimpan

    def __init__(self, path: str | Path, parent=None) -> None:
        super().__init__(parent)
        self._path = Path(path)
        self.setWindowTitle(f"Edit Excel — {self._path.name}")
        self.resize(900, 600)
        self.setWindowFlag(Qt.WindowType.Window, True)  # jendela mandiri (non-modal)

        self._wb = load_workbook(self._path)  # editable (bukan read_only)

        layout = QVBoxLayout(self)
        layout.addLayout(self._bangun_toolbar())
        self._tabs = QTabWidget()
        layout.addWidget(self._tabs, 1)
        self._muat_sheet()

    # ---- UI ------------------------------------------------------------
    def _bangun_toolbar(self) -> QHBoxLayout:
        bar = QHBoxLayout()
        self.btn_tambah_baris = QPushButton("+ Baris")
        self.btn_tambah_baris.clicked.connect(self._tambah_baris)
        self.btn_hapus_baris = QPushButton("– Baris")
        self.btn_hapus_baris.clicked.connect(self._hapus_baris)
        self.btn_tambah_kolom = QPushButton("+ Kolom")
        self.btn_tambah_kolom.clicked.connect(self._tambah_kolom)
        self.btn_simpan = QPushButton("💾 Simpan")
        self.btn_simpan.setStyleSheet("background:#1a7f37; color:white; font-weight:600; padding:6px 14px;")
        self.btn_simpan.clicked.connect(self._simpan)
        for w in (self.btn_tambah_baris, self.btn_hapus_baris, self.btn_tambah_kolom):
            bar.addWidget(w)
        bar.addStretch(1)
        bar.addWidget(self.btn_simpan)
        return bar

    def _muat_sheet(self) -> None:
        for ws in self._wb.worksheets:
            baris = list(ws.iter_rows(values_only=True))
            n_baris = max(len(baris), 1)
            n_kolom = max((len(r) for r in baris), default=1)
            table = QTableWidget(n_baris, n_kolom)
            for r, row in enumerate(baris):
                for c, val in enumerate(row):
                    table.setItem(r, c, QTableWidgetItem("" if val is None else str(val)))
            self._tabs.addTab(table, ws.title)

    # ---- Aksi tabel ----------------------------------------------------
    def _table_aktif(self) -> QTableWidget | None:
        return self._tabs.currentWidget()  # type: ignore[return-value]

    def _tambah_baris(self) -> None:
        t = self._table_aktif()
        if t is not None:
            baris = t.currentRow() + 1 if t.currentRow() >= 0 else t.rowCount()
            t.insertRow(baris)

    def _hapus_baris(self) -> None:
        t = self._table_aktif()
        if t is None:
            return
        rows = sorted({idx.row() for idx in t.selectedIndexes()}, reverse=True)
        for r in rows or ([t.rowCount() - 1] if t.rowCount() else []):
            t.removeRow(r)

    def _tambah_kolom(self) -> None:
        t = self._table_aktif()
        if t is not None:
            t.insertColumn(t.columnCount())

    # ---- Simpan --------------------------------------------------------
    def _simpan(self) -> None:
        try:
            for i in range(self._tabs.count()):
                table: QTableWidget = self._tabs.widget(i)  # type: ignore[assignment]
                ws = self._wb[self._tabs.tabText(i)]
                n_baris, n_kolom = table.rowCount(), table.columnCount()
                for r in range(n_baris):
                    for c in range(n_kolom):
                        item = table.item(r, c)
                        teks = item.text() if item is not None else ""
                        ws.cell(row=r + 1, column=c + 1).value = teks or None
                # Pangkas sisa baris/kolom lama di luar ukuran tabel.
                if ws.max_row > n_baris:
                    ws.delete_rows(n_baris + 1, ws.max_row - n_baris)
                if ws.max_column > n_kolom:
                    ws.delete_cols(n_kolom + 1, ws.max_column - n_kolom)
            self._wb.save(self._path)
        except PermissionError:
            QMessageBox.critical(
                self, "Gagal menyimpan",
                "Berkas sedang dibuka di aplikasi lain (mis. Excel). Tutup dulu lalu coba lagi.",
            )
            return
        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(self, "Gagal menyimpan", str(exc))
            return
        QMessageBox.information(self, "Tersimpan", f"Perubahan disimpan ke {self._path.name}.")
        self.tersimpan.emit(str(self._path))

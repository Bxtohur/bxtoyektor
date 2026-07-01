"""Pembaca Excel dengan deteksi header otomatis (PRD F-1.1, §13).

Struktur sumber (LINK BUKTI FISIK C1-C9.xlsx) TIDAK seragam antar sheet:
- Sheet "C1-C9": kolom B = NAMA FILE, C = LINK FILE (judul kategori muncul
  sebagai baris di kolom B dengan link kosong — "merambat" ke bawah).
- Sheet "C2"/"C4"/"C5": kolom A = JUDUL, B = NAMA FILE, C = LINK.
- Sheet "C3": hanya 2 kolom — A = NAMA FILE, B = LINK.

Karena itu kolom dideteksi otomatis dari baris header, per-sheet.
Modul ini murni logika (tanpa Qt) agar mudah diuji.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .models import DocumentItem

_WS = re.compile(r"\s+")


def _norm(v: object) -> str:
    """Normalisasi sel: str, strip, dan rapikan whitespace/newline."""
    if v is None:
        return ""
    return _WS.sub(" ", str(v)).strip()


def _norm_key(v: object) -> str:
    """Bentuk uppercase untuk pencocokan header."""
    return _norm(v).upper()


def _tampak_url_atau_path(v: str) -> bool:
    """True bila nilai kolom link berupa URL atau path file yang masuk akal."""
    low = v.lower()
    return low.startswith(("http://", "https://")) or bool(
        re.search(r"[\\/]", v)
    )


def _tampak_judul_kategori(v: str) -> bool:
    """Baris judul kategori (sheet1-style): teks bermakna, bukan angka nyasar."""
    return len(v) > 3 and not v.isdigit()


@dataclass(slots=True)
class HasilImport:
    """Hasil import satu workbook."""

    items: list[DocumentItem]
    peringatan: list[str]

    @property
    def jumlah_per_sheet(self) -> dict[str, int]:
        out: dict[str, int] = {}
        for it in self.items:
            out[it.sheet] = out.get(it.sheet, 0) + 1
        return out


def _cari_header(rows: list[list[str]]) -> tuple[int, int, int, int | None] | None:
    """Cari baris header dalam ~15 baris pertama.

    Return (idx_baris_header, kol_nama, kol_link, kol_judul|None) atau None.
    """
    for i, row in enumerate(rows[:15]):
        upper = [_norm_key(c) for c in row]
        kol_nama = next((j for j, c in enumerate(upper) if "NAMA FILE" in c), None)
        if kol_nama is None:
            continue
        kol_link = next((j for j, c in enumerate(upper) if "LINK" in c), None)
        if kol_link is None:
            # fallback: kolom tepat setelah nama dianggap link
            kol_link = kol_nama + 1
        kol_judul = next((j for j, c in enumerate(upper) if "JUDUL" in c), None)
        return i, kol_nama, kol_link, kol_judul
    return None


def _parse_sheet(nama_sheet: str, rows: list[list[str]], peringatan: list[str]) -> list[DocumentItem]:
    header = _cari_header(rows)
    if header is None:
        peringatan.append(
            f"Sheet '{nama_sheet}': baris header (NAMA FILE) tidak ditemukan, dilewati."
        )
        return []

    idx, kol_nama, kol_link, kol_judul = header
    items: list[DocumentItem] = []
    kategori_aktif = ""

    def ambil(row: list[str], k: int | None) -> str:
        return row[k] if k is not None and k < len(row) else ""

    # Judul kategori kadang berada DI ATAS baris header (mis. sheet "C1-C9"
    # baris 1 = "C1. Visi...", baris 2 = "NAMA FILE"). Ambil sebagai kategori awal.
    for row in reversed(rows[:idx]):
        kandidat = ambil(row, kol_judul) or ambil(row, kol_nama)
        if kandidat and _tampak_judul_kategori(kandidat):
            kategori_aktif = kandidat
            break

    for row in rows[idx + 1 :]:
        judul = ambil(row, kol_judul)
        if judul:  # kolom JUDUL eksplisit — merambat ke bawah
            kategori_aktif = judul

        nama = ambil(row, kol_nama)
        link = ambil(row, kol_link)

        if not nama and not link:
            continue

        if nama and not link:
            # baris judul kategori gaya sheet "C1-C9" (link kosong)
            if _tampak_judul_kategori(nama):
                kategori_aktif = nama
            continue

        if link and not _tampak_url_atau_path(link):
            # link tidak valid — lewati (bisa jadi baris keterangan)
            continue

        if nama and link:
            items.append(
                DocumentItem(
                    sheet=nama_sheet,
                    kategori=kategori_aktif,
                    nama_file=nama,
                    lokasi=link,
                )
            )

    if not items:
        peringatan.append(f"Sheet '{nama_sheet}': tidak ada baris data valid ditemukan.")
    return items


def baca_excel(path: str | Path) -> HasilImport:
    """Baca semua sheet dan normalisasi ke list[DocumentItem].

    Raise FileNotFoundError bila file tidak ada; error parsing per-sheet
    dikumpulkan sebagai peringatan (tidak menggagalkan seluruh import).
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"File Excel tidak ditemukan: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    peringatan: list[str] = []
    items: list[DocumentItem] = []
    try:
        for ws in wb.worksheets:
            rows = [
                [_norm(c) for c in r]
                for r in ws.iter_rows(values_only=True)
            ]
            items.extend(_parse_sheet(ws.title, rows, peringatan))
    finally:
        wb.close()

    if not items:
        peringatan.append("Tidak ada dokumen yang berhasil dibaca dari file ini.")
    return HasilImport(items=items, peringatan=peringatan)
